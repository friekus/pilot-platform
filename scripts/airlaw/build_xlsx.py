#!/usr/bin/env python3
"""Build data/QuestionBank_AirLaw_RPL_v3.xlsx from the v3 JSON."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC  = os.path.join(ROOT, "data", "airlaw_rpl_questions_v3.json")
DST  = os.path.join(ROOT, "data", "QuestionBank_AirLaw_RPL_v3.xlsx")

def main():
    with open(SRC) as f:
        qs = json.load(f)
    wb = Workbook()
    ws = wb.active
    ws.title = "Air Law RPL v3"
    header = ["#","Subject","Level","Subtopic","Difficulty","Question","A","B","C","D","Correct","Explanation","Reference","Tags"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for i,q in enumerate(qs, 1):
        ws.append([
            i, q["subject"], q["level"], q["subtopic"], q["difficulty"], q["question"],
            q["option_a"], q["option_b"], q["option_c"], q["option_d"],
            q["correct_answer"], q["explanation"], q["reference"],
            ", ".join(q.get("tags", []))
        ])
    widths = [4, 10, 6, 32, 5, 60, 30, 30, 30, 30, 8, 70, 80, 30]
    for i,w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A'+chr(64+i-26)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(DST)
    print(f"wrote {DST}")

if __name__ == "__main__":
    main()
